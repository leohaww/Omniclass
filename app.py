# ═══════════════════════════════════════════════════════════════
#  OmniClass — app.py  (All roles fully connected)
# ═══════════════════════════════════════════════════════════════
import os
from datetime import datetime, date, time, timedelta
from functools import wraps
from dotenv import load_dotenv
load_dotenv()

from flask import (Flask, Blueprint, render_template, redirect, url_for,
                   flash, request, jsonify)
from flask_login import login_user, logout_user, login_required, current_user
from flask_wtf.csrf import CSRFProtect

from db import (db, login_manager, notify, notify_many,
                User, Institution, Course, enrollments, Schedule,
                AttendanceSession, Attendance, PermitRequest,
                Assignment, Submission, Grade,
                Notification, Announcement, AuditLog)

# init_db menyediakan init global (buat database + tabel + schema + seeder)
try:
    from init_db import init_db
except Exception:
    init_db = None


# ── App & Config ─────────────────────────────────────────────────
app = Flask(__name__)
app.config.update(
    SECRET_KEY=os.getenv('SECRET_KEY','omniclass-dev-secret-2024'),
    SQLALCHEMY_DATABASE_URI=(
        f"mysql+pymysql://{os.getenv('DB_USER','root')}:"
        f"{os.getenv('DB_PASSWORD','')}@"
        f"{os.getenv('DB_HOST','localhost')}:"
        f"{os.getenv('DB_PORT','3306')}/"
        f"{os.getenv('DB_NAME','omniclass_db')}?charset=utf8mb4"),
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    SQLALCHEMY_ENGINE_OPTIONS={'pool_recycle':300,'pool_pre_ping':True},
    WTF_CSRF_ENABLED=True,
    MAX_CONTENT_LENGTH=50*1024*1024,
    UPLOAD_FOLDER=os.path.join(os.path.dirname(__file__),'uploads'),
)

db.init_app(app)
login_manager.init_app(app)
csrf = CSRFProtect(app)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

IMAGE_EXTS = {'png','jpg','jpeg','gif','webp','bmp','svg'}
ZIP_EXTS   = {'zip','rar','7z'}
DOC_EXTS   = {'pdf','doc','docx','ppt','pptx','xls','xlsx','txt','csv'}

def file_kind(filename):
    ext = (filename.rsplit('.',1)[-1] if '.' in filename else '').lower()
    if ext in IMAGE_EXTS: return 'image'
    if ext in ZIP_EXTS:   return 'zip'
    if ext in DOC_EXTS:   return 'document'
    return 'other'

def human_size(num_bytes):
    if not num_bytes: return '—'
    for unit in ['B','KB','MB','GB']:
        if num_bytes < 1024: return f"{num_bytes:.1f} {unit}" if unit!='B' else f"{int(num_bytes)} {unit}"
        num_bytes /= 1024
    return f"{num_bytes:.1f} TB"

def save_upload(file_storage, subfolder):
    """Save an uploaded file under static/uploads/<subfolder>/ and return metadata."""
    from werkzeug.utils import secure_filename
    import uuid
    orig_name = file_storage.filename
    ext = orig_name.rsplit('.',1)[-1].lower() if '.' in orig_name else ''
    safe_base = secure_filename(orig_name) or 'file'
    stored_name = f"{uuid.uuid4().hex}_{safe_base}"
    folder = os.path.join(app.static_folder, 'uploads', subfolder)
    os.makedirs(folder, exist_ok=True)
    full_path = os.path.join(folder, stored_name)
    file_storage.save(full_path)
    size = os.path.getsize(full_path)
    return {
        'url': f"/static/uploads/{subfolder}/{stored_name}",
        'name': orig_name,
        'ext': ext,
        'size': size,
        'kind': file_kind(orig_name),
        'disk_path': full_path,
    }

# ── Schema auto-migration (adds new columns/enum values on existing DBs) ──
#
# NOTE:
# Fungsi ensure_database_exists() sekarang dipindahkan ke init_db.py.
# Di sini masih dipertahankan agar backward-compatibility, tetapi
# app.py tidak wajib memanggilnya lagi.
def ensure_database_exists():
    """Create MySQL database if it doesn't exist yet.

    This prevents startup failure when DB_NAME hasn't been created.
    """
    import pymysql

    db_user = os.getenv('DB_USER', 'root')
    db_password = os.getenv('DB_PASSWORD', '')
    db_host = os.getenv('DB_HOST', 'localhost')
    db_port = int(os.getenv('DB_PORT', '3306'))
    db_name = os.getenv('DB_NAME', 'omniclass_db')

    # Connect without selecting the target database
    conn = pymysql.connect(
        host=db_host,
        user=db_user,
        password=db_password,
        port=db_port,
        autocommit=True,
        cursorclass=pymysql.cursors.DictCursor,
    )
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT SCHEMA_NAME FROM INFORMATION_SCHEMA.SCHEMATA WHERE SCHEMA_NAME=%s",
                (db_name,),
            )
            exists = cur.fetchone() is not None
            if not exists:
                cur.execute(f"CREATE DATABASE `{db_name}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
                print(f"✅ MySQL database created: {db_name}")
    finally:
        conn.close()


def ensure_schema():
    from sqlalchemy import text, inspect as sa_inspect
    try:
        insp = sa_inspect(db.engine)
        with db.engine.begin() as conn:
            course_cols = [c['name'] for c in insp.get_columns('courses')]
            if 'program' not in course_cols:
                conn.execute(text(
                    "ALTER TABLE courses ADD COLUMN program ENUM('PPL','DM') NOT NULL DEFAULT 'PPL'"))
            user_cols = [c['name'] for c in insp.get_columns('users')]
            if 'program' not in user_cols:
                conn.execute(text(
                    "ALTER TABLE users ADD COLUMN program ENUM('PPL','DM') NULL"))
            sub_cols = [c['name'] for c in insp.get_columns('submissions')]
            if 'revision_notes' not in sub_cols:
                conn.execute(text("ALTER TABLE submissions ADD COLUMN revision_notes TEXT"))
            if 'revision_count' not in sub_cols:
                conn.execute(text("ALTER TABLE submissions ADD COLUMN revision_count INT DEFAULT 0"))
            if 'last_revision_at' not in sub_cols:
                conn.execute(text("ALTER TABLE submissions ADD COLUMN last_revision_at DATETIME NULL"))
            if 'file_size' not in sub_cols:
                conn.execute(text("ALTER TABLE submissions ADD COLUMN file_size INT NULL"))
            conn.execute(text(
                "ALTER TABLE submissions MODIFY COLUMN status "
                "ENUM('submitted','graded','returned','revision') DEFAULT 'submitted'"))
    except Exception as ex:
        print(f"⚠️  Schema check skipped/failed (non-fatal): {ex}")

# ── Jinja globals / filters ───────────────────────────────────────
@app.template_global('now')
def _now(): return datetime.utcnow()

@app.template_filter('date_id')
def _date_id(dt):
    MO={1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'Mei',6:'Jun',
        7:'Jul',8:'Agu',9:'Sep',10:'Okt',11:'Nov',12:'Des'}
    return f"{dt.day} {MO[dt.month]} {dt.year}" if dt else '—'

@app.template_filter('time_ago')
def _time_ago(dt):
    if not dt: return '—'
    d=datetime.utcnow()-dt
    if d.seconds<60: return 'Baru saja'
    if d.seconds<3600: return f"{d.seconds//60} mnt lalu"
    if d.days<1: return f"{d.seconds//3600} jam lalu"
    if d.days<30: return f"{d.days} hari lalu"
    return dt.strftime('%d %b %Y')

@app.context_processor
def inject_globals():
    count = 0
    if current_user.is_authenticated:
        count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    return dict(app_name=os.getenv('APP_NAME','OmniClass'), unread_notif_count=count)

@app.errorhandler(403)
def e403(e): return render_template('shared/403.html'), 403
@app.errorhandler(404)
def e404(e): return render_template('shared/404.html'), 404
@app.errorhandler(500)
def e500(e): return render_template('shared/500.html'), 500

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for(f'{current_user.role}.dashboard'))
    return redirect(url_for('auth.login'))

# ── Role decorators ───────────────────────────────────────────────
def role_req(*roles):
    def dec(f):
        @wraps(f)
        @login_required
        def w(*a,**k):
            if current_user.role not in roles:
                flash('Akses ditolak.','danger')
                return redirect(url_for('index'))
            return f(*a,**k)
        return w
    return dec

director_only = lambda f: role_req('director')(f)
lecturer_only = lambda f: role_req('lecturer')(f)
student_only  = lambda f: role_req('student')(f)


# ═══════════════════════════════════════════════════════════════
#  AUTH
# ═══════════════════════════════════════════════════════════════
auth = Blueprint('auth', __name__, url_prefix='/auth')

@auth.route('/login', methods=['GET','POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for(f'{current_user.role}.dashboard'))
    if request.method == 'POST':
        ident = request.form.get('identifier','').strip()
        pwd   = request.form.get('password','')
        user  = User.query.filter((User.email==ident)|(User.username==ident)).first()
        if user and user.check_password(pwd) and user.is_active:
            login_user(user, remember=request.form.get('remember')=='on')
            user.last_login = datetime.utcnow()
            user.last_ip    = request.remote_addr
            AuditLog.log('login','User',user.id)
            db.session.commit()
            return redirect(request.args.get('next') or url_for(f'{user.role}.dashboard'))
        flash('Email/username atau password salah.','danger')
    return render_template('auth/login.html')

@auth.route('/logout')
@login_required
def logout():
    AuditLog.log('logout','User',current_user.id)
    db.session.commit()
    logout_user()
    flash('Anda telah keluar.','success')
    return redirect(url_for('auth.login'))

@auth.route('/profile', methods=['GET','POST'])
@login_required
def profile():
    if request.method == 'POST':
        ft = request.form.get('form_type','info')
        if ft == 'info':
            current_user.full_name = request.form.get('full_name',current_user.full_name).strip()
            current_user.phone     = request.form.get('phone','').strip()
            # handle avatar
            av = request.files.get('avatar')
            if av and av.filename:
                import uuid
                ext = av.filename.rsplit('.',1)[-1].lower()
                if ext in {'jpg','jpeg','png','gif','webp'}:
                    fn = f"avatars/{current_user.id}_{uuid.uuid4().hex[:8]}.{ext}"
                    save_dir = os.path.join(app.static_folder,'img','avatars')
                    os.makedirs(save_dir, exist_ok=True)
                    av.save(os.path.join(app.static_folder,'img',fn))
                    current_user.avatar = fn
            AuditLog.log('profile_updated','User',current_user.id)
            db.session.commit()
            flash('Profil berhasil diperbarui!','success')
        elif ft == 'password':
            old = request.form.get('old_password','')
            new = request.form.get('new_password','')
            cnf = request.form.get('confirm_password','')
            if not current_user.check_password(old):
                flash('Password lama salah.','danger')
            elif new != cnf:
                flash('Konfirmasi password tidak cocok.','danger')
            elif len(new) < 8:
                flash('Password minimal 8 karakter.','warning')
            else:
                current_user.set_password(new)
                AuditLog.log('password_changed','User',current_user.id)
                db.session.commit()
                flash('Password berhasil diperbarui!','success')
        return redirect(url_for('auth.profile'))

    # stats per role
    uid  = current_user.id
    iid  = current_user.institution_id
    stats = {}
    if current_user.role == 'director':
        stats = dict(
            total_students =User.query.filter_by(institution_id=iid,role='student',is_active=True).count(),
            total_lecturers=User.query.filter_by(institution_id=iid,role='lecturer',is_active=True).count(),
            total_courses  =Course.query.filter_by(institution_id=iid,is_active=True).count())
    elif current_user.role == 'lecturer':
        cs = Course.query.filter_by(lecturer_id=uid,is_active=True).all()
        stu_ids = set()
        for c in cs:
            for s in c.students.all(): stu_ids.add(s.id)
        stats = dict(
            total_courses=len(cs), total_students=len(stu_ids),
            total_assignments=Assignment.query.join(Course).filter(Course.lecturer_id==uid).count())
    else:
        ts = AttendanceSession.query.join(Course).join(
            enrollments,Course.id==enrollments.c.course_id
        ).filter(enrollments.c.user_id==uid,AttendanceSession.is_open==False).count()
        pr = Attendance.query.filter_by(student_id=uid,status='hadir').count()
        gs = Grade.query.filter_by(student_id=uid).filter(Grade.gpa_points.isnot(None)).all()
        tg = sum(g.gpa_points*(g.course.credits or 3) for g in gs)
        tc = sum(g.course.credits or 3 for g in gs)
        enrolled_cnt = db.session.query(enrollments).filter_by(user_id=uid).count()
        stats = dict(
            att_rate=round(pr/ts*100,1) if ts else 0,
            gpa=round(tg/tc,2) if tc else 0,
            courses=enrolled_cnt)

    recent_logs = AuditLog.query.filter_by(user_id=uid).order_by(AuditLog.created_at.desc()).limit(10).all()
    return render_template('auth/profile.html', user=current_user, stats=stats, recent_logs=recent_logs)

@auth.route('/notifications')
@login_required
def notifications():
    ns = Notification.query.filter_by(user_id=current_user.id)\
         .order_by(Notification.created_at.desc()).limit(50).all()
    for n in ns:
        if not n.is_read: n.mark_read()
    db.session.commit()
    return render_template('auth/notifications.html', notifications=ns)

app.register_blueprint(auth)


# ═══════════════════════════════════════════════════════════════
#  DIRECTOR
# ═══════════════════════════════════════════════════════════════
director = Blueprint('director', __name__, url_prefix='/director')

@director.route('/dashboard')
@director_only
def dashboard():
    iid   = current_user.institution_id
    today = date.today()
    total_students  = User.query.filter_by(institution_id=iid,role='student',is_active=True).count()
    total_lecturers = User.query.filter_by(institution_id=iid,role='lecturer',is_active=True).count()
    total_courses   = Course.query.filter_by(institution_id=iid,is_active=True).count()
    today_sessions  = AttendanceSession.query.join(Course).filter(
        Course.institution_id==iid, AttendanceSession.session_date==today).count()
    week_start = today - timedelta(days=today.weekday())
    sids = [s.id for s in AttendanceSession.query.join(Course).filter(
        Course.institution_id==iid, AttendanceSession.session_date>=week_start).all()]
    if sids:
        tr = Attendance.query.filter(Attendance.session_id.in_(sids)).count()
        pr = Attendance.query.filter(Attendance.session_id.in_(sids),Attendance.status=='hadir').count()
        att_rate = round(pr/tr*100,1) if tr else 0
    else:
        att_rate = 0
    at_risk       = _get_at_risk(iid)
    announcements = Announcement.query.filter_by(institution_id=iid,is_published=True)\
                    .order_by(Announcement.created_at.desc()).limit(5).all()
    recent_logs   = AuditLog.query.filter_by(institution_id=iid)\
                    .order_by(AuditLog.created_at.desc()).limit(10).all()
    chart_data    = _weekly_chart(iid)
    # Recent activity across all roles
    recent_sessions = AttendanceSession.query.join(Course).filter(
        Course.institution_id==iid).order_by(AttendanceSession.created_at.desc()).limit(5).all()
    recent_submissions = Submission.query.join(Assignment).join(Course).filter(
        Course.institution_id==iid).order_by(Submission.submitted_at.desc()).limit(5).all()
    pending_permits = PermitRequest.query.join(Course).filter(
        Course.institution_id==iid, PermitRequest.status=='pending').count()
    return render_template('director/dashboard.html',
        total_students=total_students, total_lecturers=total_lecturers,
        total_courses=total_courses, today_sessions=today_sessions,
        att_rate=att_rate, at_risk=at_risk, announcements=announcements,
        recent_logs=recent_logs, chart_data=chart_data,
        recent_sessions=recent_sessions, recent_submissions=recent_submissions,
        pending_permits=pending_permits)

@director.route('/users')
@director_only
def users():
    iid  = current_user.institution_id
    role = request.args.get('role','all')
    q    = request.args.get('q','')
    page = request.args.get('page',1,type=int)
    query= User.query.filter_by(institution_id=iid)
    if role != 'all': query = query.filter_by(role=role)
    if q: query = query.filter(
        (User.full_name.ilike(f'%{q}%'))|(User.email.ilike(f'%{q}%'))|(User.nip_nim.ilike(f'%{q}%')))
    pagination = query.order_by(User.created_at.desc()).paginate(page=page,per_page=20,error_out=False)
    return render_template('director/users.html', pagination=pagination, role_filter=role, search=q)

@director.route('/users/add', methods=['GET','POST'])
@director_only
def add_user():
    if request.method == 'POST':
        email = request.form.get('email','').strip().lower()
        uname = request.form.get('username','').strip()
        role  = request.form.get('role','student')
        if User.query.filter_by(email=email).first():
            flash('Email sudah terdaftar.','danger')
            return redirect(url_for('director.add_user'))
        if User.query.filter_by(username=uname).first():
            flash('Username sudah digunakan.','danger')
            return redirect(url_for('director.add_user'))
        program = request.form.get('program') if role in ('lecturer','student') else None
        if program not in ('PPL','DM'):
            program = None
        u = User(institution_id=current_user.institution_id,
                 full_name=request.form.get('full_name','').strip(),
                 email=email, username=uname,
                 role=role, program=program,
                 nip_nim=request.form.get('nip_nim','').strip(),
                 phone=request.form.get('phone','').strip(),
                 is_active=True, is_verified=True)
        u.set_password(request.form.get('password','password123'))
        db.session.add(u)
        db.session.flush()
        # Notify the new user
        prog_txt = f' — Kelas {u.program}' if u.program else ''
        notify(u.id, '🎉 Selamat Datang di OmniClass!',
               f'Akun Anda ({u.role}{prog_txt}) telah dibuat oleh {current_user.full_name}.',
               'success', 'system')
        AuditLog.log('user_created','User',u.id,message=email)
        db.session.commit()
        flash(f'Pengguna {u.full_name} berhasil ditambahkan.','success')
        return redirect(url_for('director.users'))
    return render_template('director/add_user.html')

@director.route('/users/<int:uid>/toggle', methods=['POST'])
@director_only
def toggle_user(uid):
    u = User.query.filter_by(id=uid,institution_id=current_user.institution_id).first_or_404()
    u.is_active = not u.is_active
    AuditLog.log('user_toggle','User',uid)
    db.session.commit()
    flash(f"Pengguna {u.full_name} {'diaktifkan' if u.is_active else 'dinonaktifkan'}.",'success')
    return redirect(url_for('director.users'))

@director.route('/users/<int:uid>/detail')
@director_only
def user_detail(uid):
    u = User.query.filter_by(id=uid,institution_id=current_user.institution_id).first_or_404()
    stats = {}
    if u.role == 'lecturer':
        courses = Course.query.filter_by(lecturer_id=u.id).all()
        stats = {
            'total_courses': len(courses),
            'total_students': sum(c.student_count for c in courses),
            'courses': [{'name': c.name, 'code': c.code, 'program': c.program,
                         'students': c.student_count} for c in courses],
        }
    elif u.role == 'student':
        courses = db.session.query(Course).join(
            enrollments, Course.id==enrollments.c.course_id).filter(enrollments.c.user_id==u.id).all()
        grades = Grade.query.filter_by(student_id=u.id).filter(Grade.gpa_points.isnot(None)).all()
        gpa = round(sum(g.gpa_points for g in grades)/len(grades),2) if grades else None
        submissions = Submission.query.filter_by(student_id=u.id).count()
        stats = {
            'total_courses': len(courses),
            'gpa': gpa,
            'total_submissions': submissions,
            'courses': [{'name': c.name, 'code': c.code, 'program': c.program} for c in courses],
        }
    inst = Institution.query.get(u.institution_id)
    return jsonify({
        'success': True,
        'id': u.id, 'full_name': u.full_name, 'email': u.email, 'username': u.username,
        'role': u.role, 'program': u.program, 'nip_nim': u.nip_nim, 'phone': u.phone,
        'avatar': u.avatar_url, 'is_active': u.is_active,
        'institution': inst.name if inst else '—',
        'created_at': u.created_at.strftime('%d %B %Y') if u.created_at else '—',
        'last_login': u.last_login.strftime('%d %B %Y, %H:%M') if u.last_login else 'Belum pernah login',
        'stats': stats,
    })

@director.route('/users/<int:uid>/edit', methods=['POST'])
@director_only
def edit_user(uid):
    u = User.query.filter_by(id=uid,institution_id=current_user.institution_id).first_or_404()
    email = request.form.get('email','').strip().lower()
    uname = request.form.get('username','').strip()
    if email != u.email and User.query.filter_by(email=email).first():
        flash('Email sudah digunakan pengguna lain.','danger')
        return redirect(url_for('director.users'))
    if uname != u.username and User.query.filter_by(username=uname).first():
        flash('Username sudah digunakan pengguna lain.','danger')
        return redirect(url_for('director.users'))
    u.full_name = request.form.get('full_name', u.full_name).strip()
    u.email     = email or u.email
    u.username  = uname or u.username
    u.phone     = request.form.get('phone', u.phone or '').strip()
    u.nip_nim   = request.form.get('nip_nim', u.nip_nim or '').strip()
    new_role = request.form.get('role', u.role)
    if new_role in ('director','lecturer','student'):
        u.role = new_role
    prog = request.form.get('program')
    u.program = prog if prog in ('PPL','DM') else (None if u.role == 'director' else u.program)
    new_password = request.form.get('new_password','').strip()
    pw_changed = False
    if new_password:
        u.set_password(new_password)
        pw_changed = True
    AuditLog.log('user_edited','User',uid,message='password_changed' if pw_changed else 'profile_updated')
    db.session.commit()
    if pw_changed:
        notify(u.id, '🔒 Password Diperbarui',
               f'Password akun Anda telah diperbarui oleh {current_user.full_name}.',
               'warning','system')
        db.session.commit()
    flash(f'Data {u.full_name} berhasil diperbarui.' + (' Password baru telah diset.' if pw_changed else ''),'success')
    return redirect(url_for('director.users'))

@director.route('/users/<int:uid>/delete', methods=['POST'])
@director_only
def delete_user(uid):
    u = User.query.filter_by(id=uid,institution_id=current_user.institution_id).first_or_404()
    if u.id == current_user.id:
        flash('Anda tidak bisa menghapus akun Anda sendiri.','danger')
        return redirect(url_for('director.users'))

    # Hard blockers — data that can't be safely orphaned without explicit reassignment
    blockers = []
    if Course.query.filter_by(lecturer_id=u.id).count() > 0:
        blockers.append('masih mengampu kelas')
    if Assignment.query.filter_by(lecturer_id=u.id).count() > 0:
        blockers.append('masih memiliki tugas yang ia buat')
    if AttendanceSession.query.filter_by(lecturer_id=u.id).count() > 0:
        blockers.append('masih memiliki sesi absensi yang ia buat')
    if Announcement.query.filter_by(author_id=u.id).count() > 0:
        blockers.append('masih memiliki pengumuman yang ia buat')
    if blockers:
        flash(f'{u.full_name} tidak dapat dihapus karena {", dan ".join(blockers)}. '
              f'Alihkan/hapus data tersebut terlebih dahulu.', 'danger')
        return redirect(url_for('director.users'))

    name = u.full_name
    try:
        # Soft references (nullable FKs) — clear so the row can be safely removed
        Attendance.query.filter_by(verified_by=u.id).update({'verified_by': None})
        PermitRequest.query.filter_by(reviewed_by=u.id).update({'reviewed_by': None})
        Submission.query.filter_by(graded_by=u.id).update({'graded_by': None})
        Grade.query.filter_by(updated_by=u.id).update({'updated_by': None})
        AuditLog.query.filter_by(user_id=u.id).update({'user_id': None})

        if u.role == 'student':
            db.session.execute(enrollments.delete().where(enrollments.c.user_id==u.id))
            Attendance.query.filter_by(student_id=u.id).delete()
            Submission.query.filter_by(student_id=u.id).delete()
            Grade.query.filter_by(student_id=u.id).delete()
            PermitRequest.query.filter_by(student_id=u.id).delete()

        Notification.query.filter_by(user_id=u.id).delete()
        AuditLog.log('user_deleted','User',uid,message=name)
        db.session.delete(u)
        db.session.commit()
        flash(f'Pengguna {name} berhasil dihapus.','success')
    except Exception as ex:
        db.session.rollback()
        flash(f'Gagal menghapus pengguna: {ex}','danger')
    return redirect(url_for('director.users'))

# ── Director: Manage Courses (assign lecturer + enroll students) ──
@director.route('/courses')
@director_only
def courses():
    iid  = current_user.institution_id
    q    = request.args.get('q','')
    program_filter = request.args.get('program','all')
    page = request.args.get('page',1,type=int)
    query= Course.query.filter_by(institution_id=iid)
    if q: query = query.filter(
        (Course.name.ilike(f'%{q}%'))|(Course.code.ilike(f'%{q}%')))
    if program_filter in ('PPL','DM'):
        query = query.filter(Course.program==program_filter)
    pagination = query.order_by(Course.created_at.desc()).paginate(page=page,per_page=20,error_out=False)
    lecturers  = User.query.filter_by(institution_id=iid,role='lecturer',is_active=True).all()
    students   = User.query.filter_by(institution_id=iid,role='student',is_active=True).all()
    count_ppl  = Course.query.filter_by(institution_id=iid,program='PPL').count()
    count_dm   = Course.query.filter_by(institution_id=iid,program='DM').count()
    return render_template('director/courses.html',
        pagination=pagination, search=q, lecturers=lecturers, students=students,
        program_filter=program_filter, count_ppl=count_ppl, count_dm=count_dm)

@director.route('/courses/create', methods=['POST'])
@director_only
def create_course():
    iid = current_user.institution_id
    code = request.form.get('code','').strip().upper()
    if Course.query.filter_by(institution_id=iid, code=code).first():
        flash(f'Kode "{code}" sudah ada.','danger')
        return redirect(url_for('director.courses'))
    program = request.form.get('program','PPL')
    if program not in ('PPL','DM'): program = 'PPL'
    c = Course(
        institution_id=iid,
        lecturer_id=request.form.get('lecturer_id',type=int),
        program=program,
        code=code, name=request.form.get('name','').strip(),
        description=request.form.get('description','').strip(),
        credits=int(request.form.get('credits',3)),
        semester=request.form.get('semester',''),
        academic_year=request.form.get('academic_year',''),
        room=request.form.get('room',''),
        capacity=int(request.form.get('capacity',40)),
        is_active=True)
    db.session.add(c); db.session.flush()
    # Enroll selected students
    student_ids = request.form.getlist('student_ids[]')
    for sid in student_ids:
        db.session.execute(enrollments.insert().values(
            user_id=int(sid), course_id=c.id, status='active'))
    # Notify lecturer
    if c.lecturer_id:
        notify(c.lecturer_id, f'📚 Kelas Baru Ditetapkan',
               f'Anda ditetapkan sebagai dosen "{c.name}" ({c.code}).',
               'info','course', url_for('lecturer.course_detail',cid=c.id))
    # Notify enrolled students
    for sid in student_ids:
        notify(int(sid), f'📖 Terdaftar di {c.name}',
               f'Anda telah didaftarkan ke mata kuliah {c.name} ({c.code}).',
               'info','course')
    AuditLog.log('course_created','Course',c.id,message=c.name)
    db.session.commit()
    flash(f'Mata kuliah "{c.name}" berhasil dibuat.','success')
    return redirect(url_for('director.courses'))

@director.route('/courses/<int:cid>/enroll', methods=['POST'])
@director_only
def enroll_students(cid):
    c = Course.query.filter_by(id=cid,institution_id=current_user.institution_id).first_or_404()
    sids = request.form.getlist('student_ids[]')
    added = 0
    for sid in sids:
        sid = int(sid)
        existing = db.session.query(enrollments).filter_by(
            user_id=sid, course_id=cid).first()
        if not existing:
            db.session.execute(enrollments.insert().values(
                user_id=sid, course_id=cid, status='active'))
            notify(sid, f'📖 Terdaftar di {c.name}',
                   f'Anda telah didaftarkan ke {c.name} ({c.code}).',
                   'info','course')
            added += 1
    if added:
        # Notify lecturer about new students
        notify(c.lecturer_id, f'👥 {added} Mahasiswa Baru',
               f'{added} mahasiswa baru ditambahkan ke kelas {c.name}.',
               'info','course')
    AuditLog.log('students_enrolled','Course',cid,message=f'{added} added')
    db.session.commit()
    flash(f'{added} mahasiswa berhasil didaftarkan ke {c.name}.','success')
    return redirect(url_for('director.courses'))

@director.route('/courses/<int:cid>/unenroll', methods=['POST'])
@director_only
def unenroll_student(cid):
    c   = Course.query.filter_by(id=cid,institution_id=current_user.institution_id).first_or_404()
    sid = request.form.get('student_id',type=int)
    db.session.execute(
        enrollments.delete().where(
            enrollments.c.user_id==sid,
            enrollments.c.course_id==cid))
    notify(sid, f'ℹ️ Dikeluarkan dari {c.name}',
           f'Anda telah dikeluarkan dari mata kuliah {c.name}.',
           'warning','course')
    db.session.commit()
    flash('Mahasiswa berhasil dikeluarkan.','success')
    return redirect(url_for('director.courses'))

@director.route('/courses/<int:cid>/assign-lecturer', methods=['POST'])
@director_only
def assign_lecturer(cid):
    c = Course.query.filter_by(id=cid,institution_id=current_user.institution_id).first_or_404()
    new_lid = request.form.get('lecturer_id',type=int)
    old_lid = c.lecturer_id
    c.lecturer_id = new_lid
    if old_lid and old_lid != new_lid:
        notify(old_lid, f'ℹ️ Kelas Dipindahkan',
               f'Anda tidak lagi mengajar {c.name}.','warning','course')
    if new_lid:
        notify(new_lid, f'📚 Kelas Ditetapkan',
               f'Anda ditetapkan mengajar {c.name} ({c.code}).',
               'info','course', url_for('lecturer.course_detail',cid=cid))
    AuditLog.log('lecturer_assigned','Course',cid)
    db.session.commit()
    flash('Dosen berhasil ditetapkan.','success')
    return redirect(url_for('director.courses'))


@director.route('/courses/<int:cid>/edit', methods=['POST'])
@director_only
def edit_course(cid):
    c = Course.query.filter_by(id=cid,institution_id=current_user.institution_id).first_or_404()

    c.code           = request.form.get('code',c.code).strip().upper()
    c.name           = request.form.get('name',c.name).strip()
    c.description    = request.form.get('description',c.description or '').strip()
    prog = request.form.get('program')
    if prog in ('PPL','DM'):
        c.program = prog
    c.credits        = int(request.form.get('credits',c.credits or 3))
    c.semester       = request.form.get('semester',c.semester or '').strip()
    c.academic_year  = request.form.get('academic_year',c.academic_year or '').strip()
    c.room           = request.form.get('room',c.room or '').strip()
    c.capacity       = int(request.form.get('capacity',c.capacity or 40))

    new_lid = request.form.get('lecturer_id', type=int)
    if new_lid:
        old_lid = c.lecturer_id
        c.lecturer_id = new_lid
        if old_lid and old_lid != new_lid:
            notify(old_lid, f'ℹ️ Kelas Informasi Diubah',
                   f'Data kelas "{c.name}" telah diperbarui oleh director.',
                   'warning','course')
            notify(new_lid, f'📚 Data Kelas Diperbarui',
                   f'Anda mengajar "{c.name}" ({c.code}).',
                   'info','course')

    AuditLog.log('course_edited','Course',cid,message=c.name)
    db.session.commit()
    flash(f'Kelas "{c.name}" berhasil diperbarui.','success')
    return redirect(url_for('director.courses'))


@director.route('/courses/<int:cid>/toggle', methods=['POST'])
@director_only
def toggle_course_director(cid):
    c = Course.query.filter_by(id=cid,institution_id=current_user.institution_id).first_or_404()
    c.is_active = not c.is_active
    AuditLog.log('course_toggled','Course',cid,message=c.name)
    db.session.commit()
    flash(f'Kelas "{c.name}" sekarang {"Aktif" if c.is_active else "Nonaktif"}.','success')
    return redirect(url_for('director.courses'))


@director.route('/courses/<int:cid>/delete', methods=['POST'])
@director_only
def delete_course_director(cid):
    c = Course.query.filter_by(id=cid,institution_id=current_user.institution_id).first_or_404()
    if c.student_count > 0:
        flash(f'Tidak dapat hapus — masih ada {c.student_count} mahasiswa terdaftar.','danger')
        return redirect(url_for('director.courses'))

    name = c.name
    db.session.delete(c)
    AuditLog.log('course_deleted','Course',cid,message=name)
    db.session.commit()
    flash(f'Kelas "{name}" telah dihapus.','success')
    return redirect(url_for('director.courses'))


@director.route('/analytics')
@director_only
def analytics():
    from sqlalchemy import func
    iid = current_user.institution_id
    grade_dist  = db.session.query(Grade.letter_grade,func.count(Grade.id))\
                  .join(Course).filter(Course.institution_id==iid)\
                  .group_by(Grade.letter_grade).all()
    course_att  = _course_att_stats(iid)
    at_risk     = _get_at_risk(iid, detailed=True)
    monthly     = _monthly_trend(iid)
    return render_template('director/analytics.html',
        grade_dist=grade_dist, course_att=course_att,
        at_risk=at_risk, monthly=monthly)

@director.route('/reports')
@director_only
def reports():
    cs = Course.query.filter_by(institution_id=current_user.institution_id,is_active=True).all()
    return render_template('director/reports.html', courses=cs)


def _report_filters():
    iid   = current_user.institution_id
    cid   = request.values.get('course_id', type=int)
    prog  = request.values.get('program', '')
    dfrom = request.values.get('date_from', '')
    dto   = request.values.get('date_to', '')
    return iid, cid, prog if prog in ('PPL','DM') else None, dfrom, dto


def _report_attendance(iid, cid, prog, dfrom, dto):
    q = db.session.query(Attendance, AttendanceSession, Course, User).join(
        AttendanceSession, Attendance.session_id==AttendanceSession.id
    ).join(Course, Attendance.course_id==Course.id
    ).join(User, Attendance.student_id==User.id
    ).filter(Course.institution_id==iid)
    if cid: q = q.filter(Course.id==cid)
    if prog: q = q.filter(Course.program==prog)
    if dfrom: q = q.filter(AttendanceSession.session_date >= datetime.strptime(dfrom,'%Y-%m-%d').date())
    if dto:   q = q.filter(AttendanceSession.session_date <= datetime.strptime(dto,'%Y-%m-%d').date())
    rows = q.order_by(AttendanceSession.session_date.desc()).all()
    out = []
    for att, sess, course, stu in rows:
        out.append({
            'Tanggal': sess.session_date.strftime('%d-%m-%Y'),
            'Pertemuan': f"#{sess.meeting_number}",
            'Program': course.program,
            'Kode Kelas': course.code,
            'Mata Kuliah': course.name,
            'Dosen': course.lecturer.full_name if course.lecturer else '—',
            'Mahasiswa': stu.full_name,
            'NIM': stu.nip_nim or stu.username,
            'Status': att.status.title(),
            'Jam Check-in': att.check_in_time.strftime('%H:%M') if att.check_in_time else '—',
            'Metode': att.method_used or '—',
        })
    return out


def _report_grades(iid, cid, prog, dfrom, dto):
    q = db.session.query(Grade, Course, User).join(
        Course, Grade.course_id==Course.id
    ).join(User, Grade.student_id==User.id
    ).filter(Course.institution_id==iid)
    if cid: q = q.filter(Course.id==cid)
    if prog: q = q.filter(Course.program==prog)
    if dfrom: q = q.filter(Grade.updated_at >= datetime.strptime(dfrom,'%Y-%m-%d'))
    if dto:   q = q.filter(Grade.updated_at <= datetime.strptime(dto,'%Y-%m-%d')+timedelta(days=1))
    rows = q.order_by(Course.code, User.full_name).all()
    out = []
    for g, course, stu in rows:
        out.append({
            'Program': course.program,
            'Kode Kelas': course.code,
            'Mata Kuliah': course.name,
            'Dosen': course.lecturer.full_name if course.lecturer else '—',
            'Mahasiswa': stu.full_name,
            'NIM': stu.nip_nim or stu.username,
            'Tugas': g.assignment_score,
            'Kuis': g.quiz_score,
            'UTS': g.midterm_score,
            'UAS': g.final_score,
            'Kehadiran': g.attendance_score,
            'Total': g.total_score,
            'Huruf': g.letter_grade or '—',
            'Bobot GPA': g.gpa_points,
        })
    return out


def _report_at_risk(iid, prog):
    students_q = User.query.filter_by(institution_id=iid,role='student',is_active=True)
    out = []
    for s in students_q.all():
        courses_q = db.session.query(Course).join(
            enrollments, Course.id==enrollments.c.course_id
        ).filter(enrollments.c.user_id==s.id, Course.institution_id==iid)
        if prog: courses_q = courses_q.filter(Course.program==prog)
        courses = courses_q.all()
        if not courses: continue
        for c in courses:
            recs = Attendance.query.filter_by(student_id=s.id, course_id=c.id).all()
            rate = round(sum(1 for r in recs if r.status=='hadir')/len(recs)*100,1) if recs else 100
            g = Grade.query.filter_by(student_id=s.id, course_id=c.id).first()
            reasons = []
            if rate < 75: reasons.append(f'Kehadiran {rate}%')
            if g and g.gpa_points is not None and g.gpa_points < 2.0:
                reasons.append(f'Nilai rendah ({g.letter_grade})')
            if reasons:
                out.append({
                    'Program': c.program,
                    'Kode Kelas': c.code,
                    'Mata Kuliah': c.name,
                    'Dosen': c.lecturer.full_name if c.lecturer else '—',
                    'Mahasiswa': s.full_name,
                    'NIM': s.nip_nim or s.username,
                    'Tingkat Kehadiran (%)': rate,
                    'Nilai Huruf': g.letter_grade if g else '—',
                    'Alasan Risiko': ', '.join(reasons),
                })
    return sorted(out, key=lambda r: r['Tingkat Kehadiran (%)'])


def _report_summary(iid, prog):
    q = Course.query.filter_by(institution_id=iid, is_active=True)
    if prog: q = q.filter(Course.program==prog)
    out = []
    for c in q.all():
        students = c.students.all()
        total_sessions = c.sessions.count()
        total_att = Attendance.query.filter_by(course_id=c.id).count()
        present   = Attendance.query.filter_by(course_id=c.id, status='hadir').count()
        att_rate  = round(present/total_att*100,1) if total_att else 0
        grades    = Grade.query.filter_by(course_id=c.id).filter(Grade.total_score.isnot(None)).all()
        avg_score = round(sum(g.total_score for g in grades)/len(grades),1) if grades else 0
        out.append({
            'Program': c.program,
            'Kode Kelas': c.code,
            'Mata Kuliah': c.name,
            'Dosen': c.lecturer.full_name if c.lecturer else '—',
            'Jumlah Mahasiswa': len(students),
            'Total Sesi': total_sessions,
            'Total Tugas': c.assignments.count(),
            'Rata-rata Kehadiran (%)': att_rate,
            'Rata-rata Nilai': avg_score,
        })
    return out


_REPORT_BUILDERS = {
    'attendance': lambda iid,cid,prog,dfrom,dto: _report_attendance(iid,cid,prog,dfrom,dto),
    'grades':     lambda iid,cid,prog,dfrom,dto: _report_grades(iid,cid,prog,dfrom,dto),
    'at_risk':    lambda iid,cid,prog,dfrom,dto: _report_at_risk(iid,prog),
    'summary':    lambda iid,cid,prog,dfrom,dto: _report_summary(iid,prog),
}

_REPORT_TITLES = {
    'attendance': 'Laporan Kehadiran',
    'grades': 'Laporan Nilai',
    'at_risk': 'Mahasiswa Berisiko',
    'summary': 'Rekapitulasi Semester',
}

@director.route('/reports/preview', methods=['POST'])
@director_only
def reports_preview():
    rtype = request.values.get('type','attendance')
    if rtype not in _REPORT_BUILDERS:
        return jsonify({'success': False, 'message': 'Jenis laporan tidak dikenal.'})
    iid, cid, prog, dfrom, dto = _report_filters()
    try:
        rows = _REPORT_BUILDERS[rtype](iid, cid, prog, dfrom, dto)
    except Exception as ex:
        return jsonify({'success': False, 'message': f'Gagal memuat data: {ex}'})
    columns = list(rows[0].keys()) if rows else []
    return jsonify({
        'success': True, 'title': _REPORT_TITLES[rtype],
        'columns': columns, 'rows': rows[:200], 'total': len(rows)
    })

@director.route('/reports/export')
@director_only
def reports_export():
    from io import BytesIO
    rtype  = request.args.get('type','attendance')
    fmt    = request.args.get('format','xlsx')
    if rtype not in _REPORT_BUILDERS:
        flash('Jenis laporan tidak dikenal.','danger')
        return redirect(url_for('director.reports'))
    iid, cid, prog, dfrom, dto = _report_filters()
    rows  = _REPORT_BUILDERS[rtype](iid, cid, prog, dfrom, dto)
    title = _REPORT_TITLES[rtype]
    columns = list(rows[0].keys()) if rows else ['Info']
    if not rows: rows = [{'Info': 'Tidak ada data untuk filter yang dipilih.'}]

    AuditLog.log('report_exported', message=f'{title} ({fmt})')
    db.session.commit()

    if fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook(); ws = wb.active; ws.title = title[:31]
        ws.append([title]); ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=max(len(columns),1))
        ws['A1'].font = Font(size=14, bold=True, color='4F46E5')
        ws.append([f"Diekspor: {datetime.utcnow().strftime('%d %B %Y %H:%M')} UTC — Direktur: {current_user.full_name}"])
        ws.append([])
        header_row = ws.max_row + 1
        ws.append(columns)
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='4F46E5')
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            ws.append([row.get(c,'') for c in columns])
        for i, col in enumerate(columns, start=1):
            width = max(12, min(35, len(str(col))+4))
            ws.column_dimensions[ws.cell(row=header_row,column=i).column_letter].width = width
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        from flask import send_file
        fname = f"{rtype}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                          mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        # Printable HTML — user can "Save as PDF" via browser print dialog
        return render_template('director/report_print.html',
            title=title, columns=columns, rows=rows,
            generated_at=datetime.utcnow(), director=current_user)

@director.route('/announcements', methods=['GET','POST'])
@director_only
def announcements():
    iid = current_user.institution_id
    if request.method == 'POST':
        a = Announcement(institution_id=iid, author_id=current_user.id,
                         title=request.form.get('title'),
                         content=request.form.get('content'),
                         target=request.form.get('target','all'),
                         is_pinned=request.form.get('is_pinned')=='on')
        db.session.add(a); db.session.flush()
        # Broadcast notification based on target
        target = a.target
        if target in ('all','students'):
            uids = [u.id for u in User.query.filter_by(
                institution_id=iid,role='student',is_active=True).all()]
            notify_many(uids,f'📢 {a.title}',
                        a.content[:100]+'…' if len(a.content)>100 else a.content,
                        'info','announcement')
        if target in ('all','lecturers'):
            uids = [u.id for u in User.query.filter_by(
                institution_id=iid,role='lecturer',is_active=True).all()]
            notify_many(uids,f'📢 {a.title}',
                        a.content[:100]+'…' if len(a.content)>100 else a.content,
                        'info','announcement')
        AuditLog.log('announcement_created',message=a.title)
        db.session.commit()
        flash('Pengumuman diterbitkan dan notifikasi terkirim.','success')
        return redirect(url_for('director.announcements'))
    anns = Announcement.query.filter_by(institution_id=iid)\
           .order_by(Announcement.created_at.desc()).all()
    return render_template('director/announcements.html', announcements=anns)

@director.route('/audit-logs')
@director_only
def audit_logs():
    page = request.args.get('page',1,type=int)
    logs = AuditLog.query.filter_by(institution_id=current_user.institution_id)\
           .order_by(AuditLog.created_at.desc())\
           .paginate(page=page,per_page=30,error_out=False)
    return render_template('director/audit_logs.html', logs=logs)

@director.route('/settings', methods=['GET','POST'])
@director_only
def settings():
    import uuid
    inst = Institution.query.get(current_user.institution_id)
    if request.method == 'POST':
        ft = request.form.get('form_type','info')
        if ft == 'logo':
            lf = request.files.get('logo')
            if lf and lf.filename:
                ext = lf.filename.rsplit('.',1)[-1].lower()
                if ext in {'jpg','jpeg','png','gif','webp','svg'}:
                    fn = f"logos/logo_{inst.id}_{uuid.uuid4().hex[:8]}.{ext}"
                    save_dir = os.path.join(app.static_folder,'img','logos')
                    os.makedirs(save_dir,exist_ok=True)
                    lf.save(os.path.join(app.static_folder,'img',fn))
                    inst.logo = fn
                    db.session.commit()
                    flash('Logo berhasil diperbarui!','success')
                else:
                    flash('Format tidak didukung.','danger')
        else:
            inst.name=request.form.get('name',inst.name).strip()
            inst.phone=request.form.get('phone',inst.phone or '').strip()
            inst.email=request.form.get('email',inst.email or '').strip()
            inst.address=request.form.get('address',inst.address or '').strip()
            inst.website=request.form.get('website',inst.website or '').strip()
            AuditLog.log('institution_updated','Institution',inst.id)
            db.session.commit()
            flash('Pengaturan disimpan!','success')
        return redirect(url_for('director.settings'))
    active_users = User.query.filter_by(institution_id=inst.id,is_active=True).count()
    return render_template('director/settings.html', institution=inst, active_users=active_users)

app.register_blueprint(director)


# ═══════════════════════════════════════════════════════════════
#  LECTURER
# ═══════════════════════════════════════════════════════════════
lecturer = Blueprint('lecturer', __name__, url_prefix='/lecturer')

@lecturer.route('/dashboard')
@lecturer_only
def dashboard():
    my_courses = Course.query.filter_by(lecturer_id=current_user.id,is_active=True).all()
    today = date.today()
    day_num = today.weekday()
    today_schedules = []
    for c in my_courses:
        for sch in c.schedules:
            if sch.day_of_week == day_num and sch.is_active:
                sess = AttendanceSession.query.filter_by(course_id=c.id,session_date=today).first()
                today_schedules.append({'course':c,'schedule':sch,'session':sess})
    cids = [c.id for c in my_courses]
    pending_permits = PermitRequest.query.filter(
        PermitRequest.course_id.in_(cids),PermitRequest.status=='pending').count() if cids else 0
    ungraded = Submission.query.join(Assignment).join(Course).filter(
        Course.lecturer_id==current_user.id, Submission.score.is_(None)).count()
    upcoming = Assignment.query.join(Course).filter(
        Course.lecturer_id==current_user.id,
        Assignment.due_date>=datetime.utcnow(),
        Assignment.is_published==True
    ).order_by(Assignment.due_date).limit(5).all()
    # Recent activity by students
    recent_subs = Submission.query.join(Assignment).join(Course).filter(
        Course.lecturer_id==current_user.id
    ).order_by(Submission.submitted_at.desc()).limit(5).all()
    recent_checkins = Attendance.query.join(Course).filter(
        Course.lecturer_id==current_user.id,
        Attendance.status=='hadir',
        Attendance.check_in_time.isnot(None)
    ).order_by(Attendance.check_in_time.desc()).limit(5).all()
    return render_template('lecturer/dashboard.html',
        my_courses=my_courses, today_schedules=today_schedules,
        pending_permits=pending_permits, ungraded=ungraded,
        upcoming_assignments=upcoming,
        recent_subs=recent_subs, recent_checkins=recent_checkins)

@lecturer.route('/courses')
@lecturer_only
def courses():
    program_filter = request.args.get('program','all')
    q = Course.query.filter_by(lecturer_id=current_user.id)
    if program_filter in ('PPL','DM'):
        q = q.filter(Course.program==program_filter)
    cs = q.order_by(Course.academic_year.desc()).all()
    count_ppl = Course.query.filter_by(lecturer_id=current_user.id,program='PPL').count()
    count_dm  = Course.query.filter_by(lecturer_id=current_user.id,program='DM').count()
    return render_template('lecturer/courses.html', courses=cs,
        program_filter=program_filter, count_ppl=count_ppl, count_dm=count_dm)

@lecturer.route('/courses/create', methods=['GET','POST'])
@lecturer_only
def create_course():
    if request.method == 'POST':
        iid  = current_user.institution_id
        program = request.form.get('program','PPL')
        if program not in ('PPL','DM'): program = 'PPL'
        code = request.form.get('code','').strip().upper()
        if Course.query.filter_by(institution_id=iid,code=code).first():
            flash(f'Kode "{code}" sudah ada.','danger')
            return redirect(url_for('lecturer.courses'))
        c = Course(institution_id=iid, lecturer_id=current_user.id,
                   program=program,
                   code=code, name=request.form.get('name','').strip(),
                   description=request.form.get('description','').strip(),
                   credits=int(request.form.get('credits',3)),
                   semester=request.form.get('semester',''),
                   academic_year=request.form.get('academic_year',''),
                   room=request.form.get('room',''),
                   capacity=int(request.form.get('capacity',40)), is_active=True)
        db.session.add(c); db.session.flush()
        days=request.form.getlist('sch_day[]')
        starts=request.form.getlist('sch_start[]')
        ends=request.form.getlist('sch_end[]')
        for day,start,end in zip(days,starts,ends):
            if day and start and end:
                sh,sm=map(int,start.split(':'))
                eh,em=map(int,end.split(':'))
                db.session.add(Schedule(course_id=c.id,day_of_week=int(day),
                    start_time=time(sh,sm), end_time=time(eh,em),
                    room=c.room, is_active=True))
        # Notify director
        dirs = User.query.filter_by(institution_id=iid,role='director',is_active=True).all()
        notify_many([d.id for d in dirs],
                    f'📚 Kelas Baru: {c.name}',
                    f'Dosen {current_user.full_name} membuat kelas baru ({c.code}).',
                    'info','course')
        AuditLog.log('course_created','Course',c.id,message=c.name)
        db.session.commit()
        flash(f'Kelas "{c.name}" berhasil dibuat.','success')
        return redirect(url_for('lecturer.courses'))
    return redirect(url_for('lecturer.courses'))

@lecturer.route('/courses/<int:cid>')
@lecturer_only
def course_detail(cid):
    c = Course.query.filter_by(id=cid,lecturer_id=current_user.id).first_or_404()
    students    = c.students.all()
    sessions    = AttendanceSession.query.filter_by(course_id=cid)\
                  .order_by(AttendanceSession.session_date.desc()).limit(10).all()
    assignments = Assignment.query.filter_by(course_id=cid)\
                  .order_by(Assignment.due_date.desc()).all()
    enrolled_ids = {s.id for s in students}
    available_students = User.query.filter_by(
        institution_id=current_user.institution_id, role='student', is_active=True
    ).filter(~User.id.in_(enrolled_ids) if enrolled_ids else True).order_by(User.full_name).all()
    return render_template('lecturer/course_detail.html',
        course=c, students=students, sessions=sessions, assignments=assignments,
        available_students=available_students)

@lecturer.route('/courses/<int:cid>/students/add-existing', methods=['POST'])
@lecturer_only
def add_existing_student(cid):
    c = Course.query.filter_by(id=cid,lecturer_id=current_user.id).first_or_404()
    sid = request.form.get('student_id', type=int)
    stu = User.query.filter_by(id=sid, institution_id=current_user.institution_id, role='student').first()
    if not stu:
        flash('Mahasiswa tidak ditemukan.','danger')
        return redirect(url_for('lecturer.course_detail', cid=cid))
    existing = db.session.query(enrollments).filter_by(user_id=sid, course_id=cid).first()
    if existing:
        flash(f'{stu.full_name} sudah terdaftar di kelas ini.','warning')
        return redirect(url_for('lecturer.course_detail', cid=cid))
    db.session.execute(enrollments.insert().values(user_id=sid, course_id=cid, status='active'))
    notify(sid, f'📖 Terdaftar di {c.name}',
           f'Anda didaftarkan ke kelas {c.program} — {c.name} ({c.code}) oleh {current_user.full_name}.',
           'info','course', url_for('student.course_detail', cid=cid))
    AuditLog.log('student_added_by_lecturer','Course',cid,message=stu.email)
    db.session.commit()
    flash(f'{stu.full_name} berhasil ditambahkan ke kelas {c.name}.','success')
    return redirect(url_for('lecturer.course_detail', cid=cid))

@lecturer.route('/courses/<int:cid>/students/create', methods=['POST'])
@lecturer_only
def create_and_enroll_student(cid):
    c = Course.query.filter_by(id=cid,lecturer_id=current_user.id).first_or_404()
    email = request.form.get('email','').strip().lower()
    uname = request.form.get('username','').strip()
    if not email or not uname:
        flash('Email dan username wajib diisi.','danger')
        return redirect(url_for('lecturer.course_detail', cid=cid))
    if User.query.filter_by(email=email).first():
        flash('Email sudah terdaftar.','danger')
        return redirect(url_for('lecturer.course_detail', cid=cid))
    if User.query.filter_by(username=uname).first():
        flash('Username sudah digunakan.','danger')
        return redirect(url_for('lecturer.course_detail', cid=cid))
    stu = User(institution_id=current_user.institution_id,
               full_name=request.form.get('full_name','').strip(),
               email=email, username=uname, role='student', program=c.program,
               nip_nim=request.form.get('nip_nim','').strip(),
               phone=request.form.get('phone','').strip(),
               is_active=True, is_verified=True)
    stu.set_password(request.form.get('password','mahasiswa123'))
    db.session.add(stu); db.session.flush()
    db.session.execute(enrollments.insert().values(user_id=stu.id, course_id=cid, status='active'))
    notify(stu.id, '🎉 Selamat Datang di OmniClass!',
           f'Akun Anda dibuat oleh dosen {current_user.full_name} dan langsung terdaftar di kelas '
           f'{c.program} — {c.name} ({c.code}).', 'success','system',
           url_for('student.course_detail', cid=cid))
    AuditLog.log('student_created_by_lecturer','Course',cid,message=email)
    db.session.commit()
    flash(f'Akun mahasiswa {stu.full_name} dibuat dan langsung didaftarkan ke {c.name}.','success')
    return redirect(url_for('lecturer.course_detail', cid=cid))

@lecturer.route('/courses/<int:cid>/students/<int:sid>/remove', methods=['POST'])
@lecturer_only
def remove_student(cid, sid):
    c = Course.query.filter_by(id=cid,lecturer_id=current_user.id).first_or_404()
    db.session.execute(enrollments.delete().where(
        enrollments.c.user_id==sid, enrollments.c.course_id==cid))
    notify(sid, f'ℹ️ Dikeluarkan dari {c.name}',
           f'Anda dikeluarkan dari kelas {c.program} — {c.name} oleh dosen.', 'warning','course')
    AuditLog.log('student_removed_by_lecturer','Course',cid,message=str(sid))
    db.session.commit()
    flash('Mahasiswa dikeluarkan dari kelas.','success')
    return redirect(url_for('lecturer.course_detail', cid=cid))

@lecturer.route('/courses/<int:cid>/edit', methods=['POST'])
@lecturer_only
def edit_course(cid):
    c = Course.query.filter_by(id=cid,lecturer_id=current_user.id).first_or_404()
    c.name=request.form.get('name',c.name).strip()
    c.description=request.form.get('description',c.description or '').strip()
    prog = request.form.get('program')
    if prog in ('PPL','DM'):
        c.program = prog
    c.credits=int(request.form.get('credits',c.credits))
    c.semester=request.form.get('semester',c.semester or '').strip()
    c.academic_year=request.form.get('academic_year',c.academic_year or '').strip()
    c.room=request.form.get('room',c.room or '').strip()
    c.capacity=int(request.form.get('capacity',c.capacity or 40))
    Schedule.query.filter_by(course_id=cid).delete()
    days=request.form.getlist('sch_day[]')
    starts=request.form.getlist('sch_start[]')
    ends=request.form.getlist('sch_end[]')
    for day,start,end in zip(days,starts,ends):
        if day and start and end:
            sh,sm=map(int,start.split(':'))
            eh,em=map(int,end.split(':'))
            db.session.add(Schedule(course_id=cid,day_of_week=int(day),
                start_time=time(sh,sm),end_time=time(eh,em),
                room=c.room,is_active=True))
    # Notify enrolled students about schedule update
    notify_many([s.id for s in c.students.all()],
                f'🔄 Jadwal Diperbarui: {c.name}',
                f'Jadwal/info kelas {c.name} telah diperbarui oleh dosen.',
                'warning','course')
    AuditLog.log('course_updated','Course',cid,message=c.name)
    db.session.commit()
    flash(f'Kelas "{c.name}" berhasil diperbarui.','success')
    return redirect(url_for('lecturer.courses'))

@lecturer.route('/courses/<int:cid>/toggle', methods=['POST'])
@lecturer_only
def toggle_course(cid):
    c = Course.query.filter_by(id=cid,lecturer_id=current_user.id).first_or_404()
    c.is_active = not c.is_active
    AuditLog.log('course_toggle','Course',cid)
    db.session.commit()
    flash(f'Kelas "{c.name}" {"diaktifkan" if c.is_active else "dinonaktifkan"}.','success')
    return redirect(url_for('lecturer.courses'))

@lecturer.route('/courses/<int:cid>/delete', methods=['POST'])
@lecturer_only
def delete_course(cid):
    c = Course.query.filter_by(id=cid,lecturer_id=current_user.id).first_or_404()
    if c.student_count > 0:
        flash(f'Tidak dapat hapus — masih ada {c.student_count} mahasiswa.','danger')
        return redirect(url_for('lecturer.courses'))
    name = c.name
    db.session.delete(c)
    AuditLog.log('course_deleted','Course',cid,message=name)
    db.session.commit()
    flash(f'Kelas "{name}" dihapus.','success')
    return redirect(url_for('lecturer.courses'))

# ── Attendance ────────────────────────────────────────────────────
@lecturer.route('/attendance')
@lecturer_only
def attendance():
    cs  = Course.query.filter_by(lecturer_id=current_user.id,is_active=True).all()
    cid = request.args.get('course_id',type=int)
    sc  = None; sessions = []
    if cid:
        sc = Course.query.filter_by(id=cid,lecturer_id=current_user.id).first_or_404()
        sessions = AttendanceSession.query.filter_by(course_id=cid)\
                   .order_by(AttendanceSession.session_date.desc()).all()
    return render_template('lecturer/attendance.html',
        courses=cs, selected_course=sc, sessions=sessions)

@lecturer.route('/attendance/start', methods=['POST'])
@lecturer_only
def start_session():
    cid   = request.form.get('course_id',type=int)
    c     = Course.query.filter_by(id=cid,lecturer_id=current_user.id).first_or_404()
    today = date.today()
    exist = AttendanceSession.query.filter_by(course_id=cid,session_date=today).first()
    if exist:
        flash('Sesi absensi hari ini sudah ada.','warning')
        return redirect(url_for('lecturer.attendance_session',sid=exist.id))
    last = AttendanceSession.query.filter_by(course_id=cid)\
           .order_by(AttendanceSession.meeting_number.desc()).first()
    mn   = (last.meeting_number+1) if last else 1
    sess = AttendanceSession(
        course_id=cid, lecturer_id=current_user.id,
        session_date=today, topic=request.form.get('topic',''),
        method=request.form.get('method','qr_code'),
        meeting_number=mn, start_time=datetime.utcnow())
    sess.generate_qr_token()
    db.session.add(sess); db.session.flush()
    students = c.students.all()
    for s in students:
        db.session.add(Attendance(session_id=sess.id,course_id=cid,
                                  student_id=s.id,status='alfa'))
    # Notify all enrolled students
    notify_many([s.id for s in students],
                f'📋 Absensi Dimulai: {c.name}',
                f'Sesi #{mn} dimulai. Scan QR untuk absen.',
                'info','attendance')
    AuditLog.log('session_started','AttendanceSession',sess.id,
                 message=f'{c.name} #{mn}')
    db.session.commit()
    flash(f'Sesi #{mn} dimulai.','success')
    return redirect(url_for('lecturer.attendance_session',sid=sess.id))

@lecturer.route('/attendance/session/<int:sid>')
@lecturer_only
def attendance_session(sid):
    sess = AttendanceSession.query.filter_by(id=sid,lecturer_id=current_user.id).first_or_404()
    records = Attendance.query.filter_by(session_id=sid).all()
    return render_template('lecturer/attendance_session.html',session=sess,records=records)

@lecturer.route('/attendance/session/<int:sid>/refresh-qr', methods=['POST'])
@lecturer_only
def refresh_qr(sid):
    sess = AttendanceSession.query.filter_by(id=sid,lecturer_id=current_user.id).first_or_404()
    token = sess.generate_qr_token()
    db.session.commit()
    return jsonify({'token':token,'expires_at':sess.qr_expires_at.isoformat()})

@lecturer.route('/attendance/session/<int:sid>/update', methods=['POST'])
@lecturer_only
def update_attendance(sid):
    AttendanceSession.query.filter_by(id=sid,lecturer_id=current_user.id).first_or_404()
    att = Attendance.query.filter_by(
        session_id=sid, student_id=request.form.get('student_id',type=int)).first()
    if att:
        att.status     = request.form.get('status','alfa')
        att.verified_by= current_user.id
        db.session.commit()
    return jsonify({'success':True})

@lecturer.route('/attendance/session/<int:sid>/close', methods=['POST'])
@lecturer_only
def close_session(sid):
    sess = AttendanceSession.query.filter_by(id=sid,lecturer_id=current_user.id).first_or_404()
    sess.is_open  = False
    sess.end_time = datetime.utcnow()
    db.session.commit()
    flash('Sesi absensi ditutup.','success')
    return redirect(url_for('lecturer.attendance'))

# ── Permits ────────────────────────────────────────────────────────
@lecturer.route('/permits')
@lecturer_only
def permits():
    cids = [c.id for c in Course.query.filter_by(lecturer_id=current_user.id).all()]
    prs  = PermitRequest.query.filter(PermitRequest.course_id.in_(cids))\
           .order_by(PermitRequest.created_at.desc()).all() if cids else []
    return render_template('lecturer/permits.html', permits=prs)

@lecturer.route('/permits/<int:pid>/review', methods=['POST'])
@lecturer_only
def review_permit(pid):
    p = PermitRequest.query.get_or_404(pid)
    p.status      = 'approved' if request.form.get('action')=='approve' else 'rejected'
    p.reviewed_by = current_user.id
    p.reviewed_at = datetime.utcnow()
    p.review_notes= request.form.get('notes','')
    if p.status=='approved' and p.session_id:
        att = Attendance.query.filter_by(session_id=p.session_id,student_id=p.student_id).first()
        if att: att.status = p.type
    notify(p.student_id,
           f'{"✅" if p.status=="approved" else "❌"} Izin {p.status.title()}',
           f'Pengajuan {p.type} Anda untuk {p.course.name} telah {p.status}.',
           'success' if p.status=='approved' else 'danger','attendance')
    db.session.commit()
    flash(f'Izin {p.status}.','success')
    return redirect(url_for('lecturer.permits'))

# ── Assignments ────────────────────────────────────────────────────
@lecturer.route('/assignments')
@lecturer_only
def assignments():
    cs  = Course.query.filter_by(lecturer_id=current_user.id,is_active=True).all()
    cid = request.args.get('course_id',type=int)
    sc  = None; asgns = []
    if cid:
        sc    = Course.query.filter_by(id=cid,lecturer_id=current_user.id).first_or_404()
        asgns = Assignment.query.filter_by(course_id=cid).order_by(Assignment.due_date.desc()).all()
    return render_template('lecturer/assignments.html',
        courses=cs, selected_course=sc, assignments=asgns)

@lecturer.route('/assignments/create', methods=['GET','POST'])
@lecturer_only
def create_assignment():
    cs = Course.query.filter_by(lecturer_id=current_user.id,is_active=True).all()
    if request.method == 'POST':
        due = datetime.strptime(request.form.get('due_date'),'%Y-%m-%dT%H:%M')
        a = Assignment(
            course_id=request.form.get('course_id',type=int),
            lecturer_id=current_user.id,
            title=request.form.get('title'),
            description=request.form.get('description',''),
            type=request.form.get('type','tugas'),
            max_score=float(request.form.get('max_score',100)),
            due_date=due,
            late_penalty=float(request.form.get('late_penalty',10)),
            allow_late=request.form.get('allow_late')=='on',
            is_published=request.form.get('is_published')=='on')
        db.session.add(a); db.session.flush()
        c = Course.query.get(a.course_id)
        # Auto-open (or refresh) today's attendance QR session for this course so students
        # must check in before working on the assignment
        if a.is_published:
            today_sess = AttendanceSession.query.filter_by(
                course_id=c.id, session_date=date.today()).first()
            if not today_sess:
                meeting_no = AttendanceSession.query.filter_by(course_id=c.id).count() + 1
                today_sess = AttendanceSession(
                    course_id=c.id, lecturer_id=current_user.id, session_date=date.today(),
                    meeting_number=meeting_no, topic=f'Sesi Tugas: {a.title}', is_open=True)
                db.session.add(today_sess)
            elif not today_sess.is_open:
                today_sess.is_open = True
            today_sess.generate_qr_token()
        # Notify enrolled students
        notify_many([s.id for s in c.students.all()],
                    f'📝 Tugas Baru: {a.title}',
                    f'Tugas baru di {c.name} — deadline {due.strftime("%d %b %Y, %H:%M")}. '
                    f'Absen QR otomatis aktif, silakan check-in sebelum mengerjakan.',
                    'warning','assignment',
                    url_for('student.assignment_detail',aid=a.id))
        AuditLog.log('assignment_created','Assignment',a.id,message=a.title)
        db.session.commit()
        flash(f'Tugas "{a.title}" dibuat. Sesi absensi QR otomatis aktif untuk kelas ini.','success')
        return redirect(url_for('lecturer.assignments',course_id=a.course_id))
    return render_template('lecturer/create_assignment.html', courses=cs)

@lecturer.route('/assignments/<int:aid>/submissions')
@lecturer_only
def submissions(aid):
    a = Assignment.query.join(Course).filter(
        Assignment.id==aid,Course.lecturer_id==current_user.id).first_or_404()
    subs = Submission.query.filter_by(assignment_id=aid).all()
    return render_template('lecturer/submissions.html', assignment=a, submissions=subs)

def sync_assignment_grade(student_id, course_id):
    """Recompute Grade.assignment_score from all graded Submission scores in this course,
    then recalculate total_score/letter/gpa. Called whenever a lecturer grades/revises a submission."""
    subs = db.session.query(Submission).join(Assignment).filter(
        Assignment.course_id == course_id,
        Submission.student_id == student_id,
        Submission.score.isnot(None)
    ).all()
    g = Grade.query.filter_by(course_id=course_id, student_id=student_id).first()
    if not g:
        g = Grade(course_id=course_id, student_id=student_id)
        db.session.add(g)
    if subs:
        pct_scores = [(s.score / s.assignment.max_score * 100) for s in subs if s.assignment.max_score]
        g.assignment_score = round(sum(pct_scores) / len(pct_scores), 2) if pct_scores else 0
    else:
        g.assignment_score = 0
    g.calculate_total()
    return g

@lecturer.route('/assignments/<int:aid>/grade/<int:subid>', methods=['POST'])
@lecturer_only
def grade_submission(aid, subid):
    sub = Submission.query.filter_by(id=subid,assignment_id=aid).first_or_404()
    sub.score     = float(request.form.get('score',0))
    sub.feedback  = request.form.get('feedback','')
    sub.graded_by = current_user.id
    sub.graded_at = datetime.utcnow()
    sub.status    = 'graded'
    g = sync_assignment_grade(sub.student_id, sub.assignment.course_id)
    notify(sub.student_id,
           f'📊 Tugas Dinilai: {sub.assignment.title}',
           f'Nilai Anda: {sub.score}/{sub.assignment.max_score}. {sub.feedback[:80] if sub.feedback else ""} '
           f'Nilai tugas keseluruhan di kelas ini kini {g.assignment_score} (Total: {g.total_score}, {g.letter_grade}).',
           'success','assignment',
           url_for('student.assignment_detail',aid=aid))
    db.session.commit()
    return jsonify({'success':True,'score':sub.score,
                     'assignment_score':g.assignment_score,'total_score':g.total_score,'letter':g.letter_grade})

@lecturer.route('/assignments/<int:aid>/submissions/<int:subid>/revision', methods=['POST'])
@lecturer_only
def request_revision(aid, subid):
    a = Assignment.query.join(Course).filter(
        Assignment.id==aid, Course.lecturer_id==current_user.id).first_or_404()
    sub = Submission.query.filter_by(id=subid, assignment_id=aid).first_or_404()
    notes = request.form.get('notes','').strip()
    if not notes:
        return jsonify({'success': False, 'message': 'Catatan revisi wajib diisi.'})
    sub.status = 'revision'
    sub.revision_notes = notes
    sub.revision_count = (sub.revision_count or 0) + 1
    sub.last_revision_at = datetime.utcnow()
    sub.score = None
    sync_assignment_grade(sub.student_id, a.course_id)
    notify(sub.student_id,
           f'🔁 Perlu Revisi: {a.title}',
           f'Dosen meminta revisi tugas "{a.title}". Catatan: {notes[:120]}',
           'warning','assignment', url_for('student.assignment_detail', aid=aid))
    AuditLog.log('submission_revision_requested','Submission',sub.id)
    db.session.commit()
    return jsonify({'success': True, 'status': sub.status, 'revision_count': sub.revision_count})

@lecturer.route('/assignments/<int:aid>/submissions/<int:subid>/detail')
@lecturer_only
def submission_detail(aid, subid):
    Assignment.query.join(Course).filter(
        Assignment.id==aid, Course.lecturer_id==current_user.id).first_or_404()
    sub = Submission.query.filter_by(id=subid, assignment_id=aid).first_or_404()
    kind = file_kind(sub.file_name) if sub.file_name else None
    return jsonify({
        'success': True,
        'student_name': sub.student.full_name,
        'nim': sub.student.nip_nim or sub.student.username,
        'avatar': sub.student.avatar_url,
        'submitted_at': sub.submitted_at.strftime('%d %B %Y, %H:%M'),
        'is_late': sub.is_late,
        'text_content': sub.text_content or '',
        'file_url': sub.file_url,
        'download_url': url_for('lecturer.download_submission_file', aid=aid, subid=subid) if sub.file_url else None,
        'file_name': sub.file_name,
        'file_size': human_size(sub.file_size) if sub.file_size else None,
        'file_kind': kind,
        'status': sub.status,
        'score': sub.score,
        'feedback': sub.feedback or '',
        'revision_notes': sub.revision_notes or '',
        'revision_count': sub.revision_count or 0,
        'last_revision_at': sub.last_revision_at.strftime('%d %B %Y, %H:%M') if sub.last_revision_at else None,
        'max_score': sub.assignment.max_score,
    })

@lecturer.route('/assignments/<int:aid>/submissions/<int:subid>/download')
@lecturer_only
def download_submission_file(aid, subid):
    from flask import send_file, abort
    Assignment.query.join(Course).filter(
        Assignment.id==aid, Course.lecturer_id==current_user.id).first_or_404()
    sub = Submission.query.filter_by(id=subid, assignment_id=aid).first_or_404()
    if not sub.file_url:
        abort(404)
    disk_path = os.path.join(app.static_folder, sub.file_url.split('/static/',1)[-1])
    if not os.path.isfile(disk_path):
        abort(404)
    return send_file(disk_path, as_attachment=True,
                      download_name=sub.file_name or os.path.basename(disk_path))

@lecturer.route('/assignments/<int:aid>/submissions/<int:subid>/zip-contents')
@lecturer_only
def submission_zip_contents(aid, subid):
    Assignment.query.join(Course).filter(
        Assignment.id==aid, Course.lecturer_id==current_user.id).first_or_404()
    sub = Submission.query.filter_by(id=subid, assignment_id=aid).first_or_404()
    if not sub.file_url or file_kind(sub.file_name or '') != 'zip':
        return jsonify({'success': False, 'message': 'Bukan file ZIP.'})
    import zipfile
    disk_path = os.path.join(app.static_folder, sub.file_url.split('/static/',1)[-1])
    if not os.path.isfile(disk_path):
        return jsonify({'success': False, 'message': 'File tidak ditemukan di server.'})
    if not zipfile.is_zipfile(disk_path):
        return jsonify({'success': False,
                         'message': 'File ini bukan format ZIP standar (mungkin .rar/.7z). Gunakan tombol Unduh untuk membuka di aplikasi ekstraksi.'})
    try:
        entries = []
        with zipfile.ZipFile(disk_path) as zf:
            bad = zf.testzip()
            for info in zf.infolist():
                if info.is_dir(): continue
                entries.append({'name': info.filename, 'size': human_size(info.file_size)})
        return jsonify({'success': True, 'entries': entries[:300], 'total': len(entries)})
    except Exception as ex:
        return jsonify({'success': False, 'message': f'Gagal membaca isi ZIP: {ex}'})

# ── Grades ─────────────────────────────────────────────────────────
@lecturer.route('/grades')
@lecturer_only
def grades():
    cs  = Course.query.filter_by(lecturer_id=current_user.id,is_active=True).all()
    cid = request.args.get('course_id',type=int)
    sc  = None; gdata = []
    if cid:
        sc = Course.query.filter_by(id=cid,lecturer_id=current_user.id).first_or_404()
        for s in sc.students.all():
            g = Grade.query.filter_by(course_id=cid,student_id=s.id).first()
            if not g:
                g = Grade(course_id=cid,student_id=s.id)
                db.session.add(g)
            gdata.append({'student':s,'grade':g})
        db.session.commit()
    return render_template('lecturer/grades.html',
        courses=cs, selected_course=sc, grade_data=gdata)

@lecturer.route('/grades/update', methods=['POST'])
@lecturer_only
def update_grade():
    cid = request.form.get('course_id',type=int)
    sid = request.form.get('student_id',type=int)
    Course.query.filter_by(id=cid,lecturer_id=current_user.id).first_or_404()
    g = Grade.query.filter_by(course_id=cid,student_id=sid).first()
    if not g:
        g = Grade(course_id=cid,student_id=sid); db.session.add(g)
    g.assignment_score=float(request.form.get('assignment_score',0))
    g.quiz_score=float(request.form.get('quiz_score',0))
    g.midterm_score=float(request.form.get('midterm_score',0))
    g.final_score=float(request.form.get('final_score',0))
    g.attendance_score=float(request.form.get('attendance_score',0))
    g.updated_by=current_user.id
    g.calculate_total()
    # Notify student
    notify(sid,
           f'📈 Nilai Diperbarui',
           f'Nilai Anda di {g.course.name} diperbarui: {g.total_score} ({g.letter_grade}).',
           'info','grade', url_for('student.grades'))
    AuditLog.log('grade_updated','Grade',g.id)
    db.session.commit()
    return jsonify({'success':True,'total':g.total_score,'letter':g.letter_grade})

app.register_blueprint(lecturer)


# ═══════════════════════════════════════════════════════════════
#  STUDENT
# ═══════════════════════════════════════════════════════════════
student = Blueprint('student', __name__, url_prefix='/student')

def _enrolled(uid=None):
    uid = uid or current_user.id
    return db.session.query(Course).join(
        enrollments,Course.id==enrollments.c.course_id
    ).filter(enrollments.c.user_id==uid, Course.is_active==True).all()

@student.route('/dashboard')
@student_only
def dashboard():
    uid      = current_user.id
    enrolled = _enrolled()
    today    = date.today()
    day_num  = today.weekday()
    today_schedule = []
    for c in enrolled:
        for sch in c.schedules:
            if sch.day_of_week==day_num and sch.is_active:
                today_schedule.append({'course':c,'schedule':sch})
    ts = AttendanceSession.query.join(Course).join(
        enrollments,Course.id==enrollments.c.course_id
    ).filter(enrollments.c.user_id==uid,AttendanceSession.is_open==False).count()
    pr = Attendance.query.filter_by(student_id=uid,status='hadir').count()
    att_rate = round(pr/ts*100,1) if ts else 0
    cids = [c.id for c in enrolled]
    upcoming = Assignment.query.filter(
        Assignment.course_id.in_(cids),
        Assignment.due_date>=datetime.utcnow(),
        Assignment.is_published==True
    ).order_by(Assignment.due_date).limit(5).all() if cids else []
    sub_ids = {s.assignment_id for s in Submission.query.filter_by(student_id=uid).all()}
    unsubmitted = [a for a in upcoming if a.id not in sub_ids]
    gs = Grade.query.filter_by(student_id=uid).filter(Grade.gpa_points.isnot(None)).all()
    tg = sum(g.gpa_points*(g.course.credits or 3) for g in gs)
    tc = sum(g.course.credits or 3 for g in gs)
    gpa = round(tg/tc,2) if tc else 0
    anns = Announcement.query.filter(
        Announcement.institution_id==current_user.institution_id,
        Announcement.is_published==True,
        Announcement.target.in_(['all','students'])
    ).order_by(Announcement.is_pinned.desc(),Announcement.created_at.desc()).limit(5).all()
    open_sessions = AttendanceSession.query.join(Course).filter(
        Course.id.in_(cids), AttendanceSession.is_open==True,
        AttendanceSession.session_date==today).all() if cids else []
    return render_template('student/dashboard.html',
        enrolled=enrolled, today_schedule=today_schedule,
        att_rate=att_rate, total_sessions=ts, present_sessions=pr,
        upcoming=upcoming, unsubmitted=unsubmitted,
        gpa=gpa, announcements=anns, open_sessions=open_sessions)

@student.route('/courses')
@student_only
def courses():
    return render_template('student/courses.html', courses=_enrolled())

@student.route('/courses/<int:cid>')
@student_only
def course_detail(cid):
    en = _enrolled()
    if not any(c.id==cid for c in en):
        flash('Anda tidak terdaftar di kelas ini.','danger')
        return redirect(url_for('student.courses'))
    c = Course.query.get_or_404(cid)
    sessions = AttendanceSession.query.filter_by(course_id=cid).all()
    att_map  = {a.session_id:a for a in Attendance.query.filter_by(
        student_id=current_user.id,course_id=cid).all()}
    att_data = [{'session':s,'record':att_map.get(s.id)} for s in sessions]
    asgns    = Assignment.query.filter_by(course_id=cid,is_published=True)\
               .order_by(Assignment.due_date).all()
    sub_ids  = {s.assignment_id for s in Submission.query.filter_by(student_id=current_user.id).all()}
    active_session = AttendanceSession.query.filter_by(
        course_id=cid, session_date=date.today(), is_open=True).first()
    already_checked_in = False
    if active_session:
        rec = Attendance.query.filter_by(session_id=active_session.id, student_id=current_user.id).first()
        already_checked_in = bool(rec and rec.status=='hadir')
    return render_template('student/course_detail.html',
        course=c, att_data=att_data, assignments=asgns, submitted_ids=sub_ids,
        active_session=active_session, already_checked_in=already_checked_in)

@student.route('/courses/<int:cid>/qr-token')
@student_only
def course_qr_token(cid):
    en = _enrolled()
    if not any(c.id==cid for c in en):
        return jsonify({'active': False})
    sess = AttendanceSession.query.filter_by(
        course_id=cid, session_date=date.today(), is_open=True).first()
    if not sess:
        return jsonify({'active': False})
    if not sess.is_qr_valid():
        sess.generate_qr_token()
        db.session.commit()
    rec = Attendance.query.filter_by(session_id=sess.id, student_id=current_user.id).first()
    return jsonify({
        'active': True,
        'token': sess.qr_token,
        'expires_at': sess.qr_expires_at.isoformat(),
        'meeting_number': sess.meeting_number,
        'topic': sess.topic or '',
        'already_checked_in': bool(rec and rec.status=='hadir'),
    })

@student.route('/attendance')
@student_only
def attendance():
    en   = _enrolled()
    cid  = request.args.get('course_id',type=int)
    sc   = None; att_data=[]; summary={}
    if cid and any(c.id==cid for c in en):
        sc = Course.query.get(cid)
        sessions = AttendanceSession.query.filter_by(course_id=cid)\
                   .order_by(AttendanceSession.session_date.desc()).all()
        att_map  = {a.session_id:a for a in Attendance.query.filter_by(
            student_id=current_user.id,course_id=cid).all()}
        att_data = [{'session':s,'record':att_map.get(s.id)} for s in sessions]
        total    = len(sessions)
        present  = sum(1 for d in att_data if d['record'] and d['record'].status=='hadir')
        izin     = sum(1 for d in att_data if d['record'] and d['record'].status in('izin','sakit'))
        alfa     = total-present-izin
        rate     = round(present/total*100,1) if total else 0
        summary  = dict(total=total,present=present,izin=izin,alfa=alfa,rate=rate)
    return render_template('student/attendance.html',
        courses=en, selected_course=sc, att_data=att_data, summary=summary)

@student.route('/attendance/checkin', methods=['POST'])
@student_only
def checkin():
    token = request.form.get('token','').strip()
    cid   = request.form.get('course_id',type=int)
    lat   = request.form.get('latitude',type=float)
    lng   = request.form.get('longitude',type=float)
    en    = _enrolled()
    if not any(c.id==cid for c in en):
        return jsonify({'success':False,'message':'Anda tidak terdaftar di kelas ini.'})
    sess = AttendanceSession.query.filter_by(qr_token=token,course_id=cid,is_open=True).first()
    if not sess:
        return jsonify({'success':False,'message':'Token tidak valid atau sesi sudah ditutup.'})
    if not sess.is_qr_valid():
        return jsonify({'success':False,'message':'Token QR kedaluwarsa. Minta refresh ke dosen.'})
    att = Attendance.query.filter_by(session_id=sess.id,student_id=current_user.id).first()
    if att and att.status=='hadir':
        return jsonify({'success':False,'message':'Anda sudah tercatat hadir.'})
    if not att:
        att = Attendance(session_id=sess.id,course_id=cid,student_id=current_user.id)
        db.session.add(att)
    att.status='hadir'; att.check_in_time=datetime.utcnow()
    att.method_used='qr_code'; att.latitude=lat; att.longitude=lng
    att.ip_address=request.remote_addr
    # Notify lecturer
    notify(sess.lecturer_id,
           f'✅ {current_user.full_name} Hadir',
           f'{current_user.full_name} baru saja absen di {sess.course.name} #{sess.meeting_number}.',
           'success','attendance')
    db.session.commit()
    return jsonify({'success':True,'message':'Kehadiran berhasil dicatat! ✅'})

@student.route('/permits/submit', methods=['GET','POST'])
@student_only
def submit_permit():
    en = _enrolled()
    if request.method == 'POST':
        cid = request.form.get('course_id',type=int)
        c   = Course.query.get(cid)
        p   = PermitRequest(
            student_id=current_user.id, course_id=cid,
            type=request.form.get('type','izin'),
            reason=request.form.get('reason'),
            request_date=datetime.strptime(request.form.get('request_date'),'%Y-%m-%d').date())
        proof = request.files.get('document')
        if proof and proof.filename:
            meta = save_upload(proof, f'permits/{current_user.id}')
            p.document_url = meta['url']
        db.session.add(p); db.session.flush()
        # Notify lecturer
        if c:
            notify(c.lecturer_id,
                   f'📋 Izin dari {current_user.full_name}',
                   f'{current_user.full_name} mengajukan {p.type} untuk {c.name}.',
                   'warning','attendance',
                   url_for('lecturer.permits'))
        AuditLog.log('permit_submitted','PermitRequest',p.id)
        db.session.commit()
        flash('Pengajuan izin terkirim ke dosen.','success')
        return redirect(url_for('student.attendance'))
    return render_template('student/submit_permit.html', courses=en)

@student.route('/assignments')
@student_only
def assignments():
    en      = _enrolled()
    cids    = [c.id for c in en]
    ftype   = request.args.get('filter','all')
    sub_ids = {s.assignment_id for s in Submission.query.filter_by(student_id=current_user.id).all()}
    query   = Assignment.query.filter(
        Assignment.course_id.in_(cids),Assignment.is_published==True
    ) if cids else Assignment.query.filter_by(id=-1)
    if ftype=='pending':
        query = query.filter(~Assignment.id.in_(sub_ids),Assignment.due_date>=datetime.utcnow())
    elif ftype=='submitted':
        query = query.filter(Assignment.id.in_(sub_ids))
    elif ftype=='overdue':
        query = query.filter(~Assignment.id.in_(sub_ids),Assignment.due_date<datetime.utcnow())
    asgns    = query.order_by(Assignment.due_date).all()
    subs_map = {s.assignment_id:s for s in Submission.query.filter_by(student_id=current_user.id).all()}
    return render_template('student/assignments.html',
        assignments=asgns, submissions_map=subs_map, filter_type=ftype)

def _assignment_attendance_gate(course_id):
    """Return (active_session, already_checked_in) for the QR check-in gate before an assignment."""
    active_session = AttendanceSession.query.filter_by(
        course_id=course_id, session_date=date.today(), is_open=True).first()
    if not active_session:
        return None, True
    rec = Attendance.query.filter_by(session_id=active_session.id, student_id=current_user.id).first()
    return active_session, bool(rec and rec.status == 'hadir')

@student.route('/assignments/<int:aid>')
@student_only
def assignment_detail(aid):
    a   = Assignment.query.get_or_404(aid)
    sub = Submission.query.filter_by(assignment_id=aid,student_id=current_user.id).first()
    active_session, checked_in = _assignment_attendance_gate(a.course_id)
    return render_template('student/assignment_detail.html', assignment=a, submission=sub,
        active_session=active_session, already_checked_in=checked_in)

@student.route('/assignments/<int:aid>/submit', methods=['POST'])
@student_only
def submit_assignment(aid):
    a = Assignment.query.get_or_404(aid)
    active_session, checked_in = _assignment_attendance_gate(a.course_id)
    if active_session and not checked_in:
        flash('Anda harus absen (scan QR) terlebih dahulu sebelum mengerjakan/mengumpulkan tugas ini.','danger')
        return redirect(url_for('student.assignment_detail',aid=aid))
    existing = Submission.query.filter_by(assignment_id=aid,student_id=current_user.id).first()
    is_late = datetime.utcnow() > a.due_date
    uploaded = request.files.get('file')
    file_meta = None
    if uploaded and uploaded.filename:
        file_meta = save_upload(uploaded, f'submissions/{aid}')

    if existing:
        if existing.status != 'revision':
            flash('Sudah dikumpulkan.','warning')
            return redirect(url_for('student.assignment_detail',aid=aid))
        # Resubmission after lecturer requested revision — text and/or file can be replaced
        existing.text_content = request.form.get('text_content','')
        if file_meta:
            existing.file_url  = file_meta['url']
            existing.file_name = file_meta['name']
            existing.file_size = file_meta['size']
        existing.submitted_at = datetime.utcnow()
        existing.is_late = is_late
        existing.status = 'submitted'
        notify(a.lecturer_id,
               f'🔁 Revisi Dikumpulkan: {a.title}',
               f'{current_user.full_name} mengumpulkan ulang (revisi ke-{existing.revision_count}) "{a.title}".',
               'info','assignment', url_for('lecturer.submissions',aid=aid))
        AuditLog.log('assignment_resubmitted','Submission',existing.id)
        db.session.commit()
        flash('Revisi tugas berhasil dikumpulkan ulang!','success')
        return redirect(url_for('student.assignment_detail',aid=aid))

    sub = Submission(assignment_id=aid,student_id=current_user.id,
                     text_content=request.form.get('text_content',''),
                     is_late=is_late)
    if file_meta:
        sub.file_url  = file_meta['url']
        sub.file_name = file_meta['name']
        sub.file_size = file_meta['size']
    db.session.add(sub); db.session.flush()
    # Notify lecturer
    notify(a.lecturer_id,
           f'📤 Pengumpulan Baru: {a.title}',
           f'{current_user.full_name} mengumpulkan "{a.title}"{"(terlambat)" if is_late else ""}.',
           'info','assignment',
           url_for('lecturer.submissions',aid=aid))
    AuditLog.log('assignment_submitted','Submission',sub.id)
    db.session.commit()
    flash('Tugas berhasil dikumpulkan!' if not is_late
          else f'Tugas dikumpulkan terlambat (penalti {a.late_penalty}%/hari).',
          'success' if not is_late else 'warning')
    return redirect(url_for('student.assignment_detail',aid=aid))

@student.route('/grades')
@student_only
def grades():
    en   = _enrolled()
    gdata=[]; tg=0; tc=0
    for c in en:
        g = Grade.query.filter_by(student_id=current_user.id,course_id=c.id).first()
        gdata.append({'course':c,'grade':g})
        if g and g.gpa_points is not None:
            tg += g.gpa_points*(c.credits or 3)
            tc += c.credits or 3
    gpa = round(tg/tc,2) if tc else 0
    return render_template('student/grades.html',
        grade_data=gdata, gpa=gpa, total_credits=tc)

app.register_blueprint(student)


# ═══════════════════════════════════════════════════════════════
#  API
# ═══════════════════════════════════════════════════════════════
api = Blueprint('api', __name__, url_prefix='/api/v1')

@api.route('/notifications')
@login_required
def get_notifications():
    ns = Notification.query.filter_by(user_id=current_user.id,is_read=False)\
         .order_by(Notification.created_at.desc()).limit(10).all()
    return jsonify([{'id':n.id,'title':n.title,'message':n.message,
                     'type':n.type,'time_ago':n.time_ago,'link':n.link} for n in ns])

@api.route('/notifications/<int:nid>/read', methods=['POST'])
@login_required
def mark_read(nid):
    n = Notification.query.filter_by(id=nid,user_id=current_user.id).first_or_404()
    n.mark_read(); db.session.commit()
    return jsonify({'success':True})

@api.route('/health')
def health(): return jsonify({'status':'ok','app':'OmniClass'})

app.register_blueprint(api)


# ═══════════════════════════════════════════════════════════════
#  ANALYTICS HELPERS
# ═══════════════════════════════════════════════════════════════
def _get_at_risk(iid, detailed=False):
    students = User.query.filter_by(institution_id=iid,role='student',is_active=True).all()
    result   = []
    for s in students:
        rs=0; reasons=[]
        recs = Attendance.query.filter_by(student_id=s.id).all()
        if recs:
            rate = sum(1 for r in recs if r.status=='hadir')/len(recs)*100
            if rate < 75: rs+=2; reasons.append(f'Kehadiran {rate:.0f}%')
        fails = [g for g in Grade.query.filter_by(student_id=s.id).all()
                 if g.gpa_points and g.gpa_points < 2.0]
        if fails: rs+=len(fails); reasons.append(f'{len(fails)} MK di bawah standar')
        if rs>0: result.append({'student':s,'risk_score':rs,'reasons':reasons})
    return sorted(result,key=lambda x:x['risk_score'],reverse=True)[:20]

def _weekly_chart(iid):
    labels=[]; hadir=[]; alfa=[]
    for i in range(6,-1,-1):
        d=date.today()-timedelta(days=i*7)
        ws=d-timedelta(days=d.weekday()); we=ws+timedelta(days=6)
        sids=[s.id for s in AttendanceSession.query.join(Course).filter(
            Course.institution_id==iid,
            AttendanceSession.session_date>=ws,
            AttendanceSession.session_date<=we).all()]
        h=Attendance.query.filter(Attendance.session_id.in_(sids),Attendance.status=='hadir').count() if sids else 0
        a=Attendance.query.filter(Attendance.session_id.in_(sids),Attendance.status=='alfa').count()  if sids else 0
        labels.append(ws.strftime('%d/%m')); hadir.append(h); alfa.append(a)
    return {'labels':labels,'hadir':hadir,'alfa':alfa}

def _course_att_stats(iid):
    cs=[]
    for c in Course.query.filter_by(institution_id=iid,is_active=True).limit(10).all():
        total=Attendance.query.filter_by(course_id=c.id).count()
        pres =Attendance.query.filter_by(course_id=c.id,status='hadir').count()
        cs.append({'course':c,'rate':round(pres/total*100,1) if total else 0,'total':total})
    return sorted(cs,key=lambda x:x['rate'])

def _monthly_trend(iid):
    res=[]
    for i in range(5,-1,-1):
        md=date.today().replace(day=1)-timedelta(days=i*30)
        ms=md.replace(day=1)
        cnt=User.query.filter(User.institution_id==iid,User.role=='student',
            User.created_at>=ms).count()
        res.append({'month':md.strftime('%b %Y'),'new_students':cnt})
    return res


if __name__ == '__main__':
    # Dev run: inisialisasi DB + tabel + schema + seeder (jika OMNICLASS_SEED=1)
    if init_db:
        init_db()
    else:
        with app.app_context():
            db.create_all()
            ensure_schema()

    print("✅ OmniClass starting...")
    app.run(debug=True, host='0.0.0.0', port=5000)
else:
    # Production (gunicorn): inisialisasi DB global dengan safety.
    # Seeder biasanya tidak dijalankan kecuali OMNICLASS_SEED=1.
    if init_db:
        try:
            init_db()
        except Exception as _e:
            print(f"⚠️  Startup init_db skipped: {_e}")
    else:
        try:
            with app.app_context():
                db.create_all()
                ensure_schema()
        except Exception as _e:
            print(f"⚠️  Startup init skipped: {_e}")



